import openpyxl
from io import BytesIO
from django.http import HttpResponse


def parse_excel(file):
    """Parse uploaded Excel file and return list of student dicts."""
    try:
        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active

        students = []
        errors = []
        headers = []
         header_row_idx = None

        # Find header row (first non-empty row)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any(row):  # Found first non-empty row
                headers = [str(cell).strip().lower() if cell else '' for cell in row]
                header_row_idx = row_idx
                break

        if not header_row_idx:
            return [], ['Excel file is empty or has no headers']

        # Process data rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            # Skip until after header row
            if row_idx <= header_row_idx:
 
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                # Header row - map column names
                headers = [str(cell).strip().lower() if cell else '' for cell in row]
                 continue

            # Skip empty rows
            if not any(row):
                continue

            # Map row data to dict using headers
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers):
                    row_data[headers[col_idx]] = str(value).strip() if value else ''

            # Extract student data - flexible header matching
            student = {
                 'roll_no': row_data.get('roll_no', row_data.get('roll no', row_data.get('rollno', row_data.get('roll', row_data.get('student roll no', ''))))),
                'prn': row_data.get('prn', row_data.get('prn no', row_data.get('prn_no', row_data.get('prn no.', '')))),
                'abc_id': row_data.get('abc_id', row_data.get('abc id', '')),
                'full_name': row_data.get('full_name', row_data.get('name', row_data.get('full name', row_data.get('student name', row_data.get('student full name', ''))))),
                'phone': row_data.get('phone', row_data.get('phone no', row_data.get('mobile', row_data.get('phone_no', row_data.get('contact number (student)', ''))))),
                'email': row_data.get('email', row_data.get('email id', row_data.get('email_id', ''))),
                'parent_name': row_data.get('parent_name', row_data.get('parent name', row_data.get('guardian name', ''))),
                'parent_phone': row_data.get('parent_phone', row_data.get('parent phone', row_data.get('parent mobile', row_data.get('parent contact number (father)', '')))),
                'birthdate': row_data.get('birthdate', row_data.get('birth date', row_data.get('dob', ''))),
                'gender': row_data.get('gender', '').strip().lower() if row_data.get('gender', '').strip() else '',
                'address': row_data.get('address', row_data.get('local address', '')),
                'permanent_address': row_data.get('permanent_address', row_data.get('permanent address', '')),
                 'roll_no': row_data.get('roll_no', row_data.get('roll no', row_data.get('rollno', row_data.get('roll', '')))),
                'prn': row_data.get('prn', row_data.get('prn no', row_data.get('prn_no', ''))),
                'full_name': row_data.get('full_name', row_data.get('name', row_data.get('full name', row_data.get('student name', '')))),
                'phone': row_data.get('phone', row_data.get('phone no', row_data.get('mobile', row_data.get('phone_no', '')))),
                'email': row_data.get('email', row_data.get('email id', row_data.get('email_id', ''))),
                'parent_name': row_data.get('parent_name', row_data.get('parent name', row_data.get('guardian name', ''))),
                'parent_phone': row_data.get('parent_phone', row_data.get('parent phone', row_data.get('parent mobile', ''))),
                'address': row_data.get('address', ''),
             }

            # Validate required fields
            if not student['roll_no']:
                errors.append(f'Row {row_idx}: Missing roll number')
                continue
            if not student['full_name']:
                errors.append(f'Row {row_idx}: Missing student name')
                continue

            # Generate PRN if missing
            if not student['prn']:
                student['prn'] = f'PRN-{student["roll_no"]}'

            students.append(student)

        wb.close()
        return students, errors

    except Exception as e:
        return [], [f'Error reading Excel file: {str(e)}']


def export_students_excel(students, file_name='students'):
    """Export student list as Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # Headers
     headers = ['Roll No', 'PRN', 'ABC ID', 'Full Name', 'Phone', 'Email',
               'Parent Phone', 'Birthdate', 'Gender', 'Local Address',
               'Permanent Address', 'Class', 'Division', 'Year', 'Status']
     headers = ['Roll No', 'PRN', 'Full Name', 'Phone', 'Email',
               'Parent Name', 'Parent Phone', 'Address', 'Class',
               'Division', 'Year', 'Status']
     ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data
    for student in students:
        ws.append([
            student.roll_no,
            student.prn,
             student.abc_id,
            student.full_name,
            student.phone,
            student.email,
            student.parent_phone,
            student.birthdate.strftime('%Y-%m-%d') if student.birthdate else '',
            student.get_gender_display(),
            student.address,
            student.permanent_address,
 
            student.full_name,
            student.phone,
            student.email,
            student.parent_name,
            student.parent_phone,
            student.address,
             student.class_name,
            student.division,
            student.year,
            student.get_status_display(),
        ])

    # Auto-width columns
    for column_cells in ws.columns:
        length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
    wb.save(response)
    return response